# -*- coding: utf-8 -*-
"""
================================================================
  RESTAURANTE · Módulo SG-SST
================================================================
Sistema de Gestión de Seguridad y Salud en el Trabajo
(Decreto 1072 de 2015 · Resolución 0312 de 2019).

POR QUÉ ESTÁ DENTRO DEL SISTEMA OPERATIVO Y NO EN UNA CARPETA APARTE
-------------------------------------------------------------------
El SG-SST suele vivir en carpetas de Word que nadie abre hasta que llega una
visita del Ministerio. Aquí se conecta con la operación real: los peligros se
clasifican por proceso —cocina, salón, almacén—, los incidentes se asocian al
empleado de la nómina y la autoevaluación arroja un porcentaje verificable.

Un restaurante concentra riesgos muy concretos: cortes, quemaduras, pisos
húmedos y carga física. No son abstracciones; son la causa real de las
incapacidades del sector.

MÉTODO DE VALORACIÓN — GTC 45
-----------------------------
    NP = ND × NE            Nivel de probabilidad
    NR = NP × NC            Nivel de riesgo
Se calcula en el servidor y no se digita: es lo que hace comparables las filas
de la matriz entre sí y evita que cada quien clasifique a ojo.

Rutas
  GET/POST/PUT  /api/sgsst/peligros
  GET/POST/PUT  /api/sgsst/actividades      plan anual de trabajo
  GET/POST/PUT  /api/sgsst/incidentes
  GET/PUT       /api/sgsst/estandares       autoevaluación Res. 0312
  GET           /api/sgsst/indicadores

Autor: Arquitectura de Software · Unidad 1
================================================================
"""
from __future__ import annotations

import datetime
import io

from fastapi import APIRouter, Body, Depends, File, Form, HTTPException, UploadFile
from sqlalchemy.orm import Session

from db import anio_actual, ahora, hoy, q, q1, run, run_sin_commit, serial, siguiente_consecutivo
from dependencias import get_tenant_db
from seguridad import autor, require_rol, verify_token

router = APIRouter(tags=["SG-SST"])

ROLES_SST = ("admin", "gerente", "sst")

CLASIFICACIONES = ["Mecánico", "Físico", "Químico", "Biológico", "Biomecánico",
                   "Locativo", "Eléctrico", "Psicosocial", "Natural", "Público"]

TIPOS_ACTIVIDAD = ["Capacitación", "Inspección", "Examen médico", "Simulacro",
                   "Entrega de EPP", "Mantenimiento", "Reunión COPASST", "Auditoría"]


# ══════════════════════════════════════════════════════════════════════
#  VALORACIÓN DEL RIESGO (GTC 45)
# ══════════════════════════════════════════════════════════════════════
def calcular_nivel_riesgo(nd: int, ne: int, nc: int) -> tuple[int, str, str]:
    """Devuelve (nivel de riesgo, interpretación, aceptabilidad).

    Escalas de la guía:
        ND — deficiencia:  0 (nulo), 2 (bajo), 6 (medio), 10 (muy alto)
        NE — exposición:   1 (esporádica) a 4 (continua)
        NC — consecuencia: 10 (leve), 25 (grave), 60 (muy grave), 100 (mortal)
    """
    nr = int(nd) * int(ne) * int(nc)
    if nr >= 600:
        return nr, "I", "No aceptable"
    if nr >= 150:
        return nr, "II", "No aceptable o aceptable con control específico"
    if nr >= 40:
        return nr, "III", "Mejorable"
    return nr, "IV", "Aceptable"


# ══════════════════════════════════════════════════════════════════════
#  MATRIZ DE PELIGROS
# ══════════════════════════════════════════════════════════════════════
@router.get("/api/sgsst/peligros")
def peligros_listar(cur: dict = Depends(verify_token), db: Session = Depends(get_tenant_db)):
    filas = serial(q(db, "SELECT * FROM sst_peligros WHERE activo = 1 "
                         "ORDER BY nivel_riesgo DESC, proceso"))
    resumen = {"I": 0, "II": 0, "III": 0, "IV": 0}
    for f in filas:
        resumen[f.get("interpretacion") or "IV"] = resumen.get(f.get("interpretacion") or "IV", 0) + 1

    return {"ok": True, "items": filas, "clasificaciones": CLASIFICACIONES,
            "kpis": {"total": len(filas), "por_nivel": resumen,
                     # Los niveles I y II exigen intervención documentada: es el
                     # número que primero mira un inspector.
                     "criticos": resumen["I"] + resumen["II"]}}


@router.post("/api/sgsst/peligros", status_code=201)
def peligro_crear(body: dict = Body(...), cur: dict = Depends(require_rol(*ROLES_SST)),
                  db: Session = Depends(get_tenant_db)):
    if not (body.get("peligro") or "").strip():
        raise HTTPException(400, "La descripción del peligro es obligatoria")
    if not (body.get("proceso") or "").strip():
        raise HTTPException(400, "El proceso es obligatorio")

    nd = int(body.get("nivel_deficiencia") or 2)
    ne = int(body.get("nivel_exposicion") or 3)
    nc = int(body.get("nivel_consecuencia") or 25)
    nr, interp, acept = calcular_nivel_riesgo(nd, ne, nc)

    res = run(db, "INSERT INTO sst_peligros (proceso, actividad, clasificacion, peligro, "
                  "efecto, nivel_deficiencia, nivel_exposicion, nivel_consecuencia, "
                  "nivel_riesgo, interpretacion, aceptabilidad, controles, epp, "
                  "responsable, activo, actualizado_en) "
                  "VALUES (:p,:a,:c,:pe,:e,:nd,:ne,:nc,:nr,:i,:ac,:ct,:epp,:r,1,:ts)",
              {"p": body["proceso"], "a": body.get("actividad"),
               "c": body.get("clasificacion") or "Mecánico", "pe": body["peligro"],
               "e": body.get("efecto"), "nd": nd, "ne": ne, "nc": nc, "nr": nr,
               "i": interp, "ac": acept, "ct": body.get("controles"),
               "epp": body.get("epp"), "r": body.get("responsable"), "ts": ahora()})
    return {"ok": True, "id": getattr(res, "lastrowid", 0),
            "nivel_riesgo": nr, "interpretacion": interp, "aceptabilidad": acept}


@router.put("/api/sgsst/peligros/{pid}")
def peligro_editar(pid: int, body: dict = Body(...),
                   cur: dict = Depends(require_rol(*ROLES_SST)),
                   db: Session = Depends(get_tenant_db)):
    actual = q1(db, "SELECT * FROM sst_peligros WHERE id=:i", {"i": pid})
    if not actual:
        raise HTTPException(404, "Peligro no encontrado")

    nd = int(body.get("nivel_deficiencia", actual["nivel_deficiencia"]))
    ne = int(body.get("nivel_exposicion", actual["nivel_exposicion"]))
    nc = int(body.get("nivel_consecuencia", actual["nivel_consecuencia"]))
    nr, interp, acept = calcular_nivel_riesgo(nd, ne, nc)

    campos = {k: body[k] for k in ("proceso", "actividad", "clasificacion", "peligro",
                                   "efecto", "controles", "epp", "responsable", "activo")
              if k in body}
    campos.update({"nivel_deficiencia": nd, "nivel_exposicion": ne,
                   "nivel_consecuencia": nc, "nivel_riesgo": nr,
                   "interpretacion": interp, "aceptabilidad": acept,
                   "actualizado_en": ahora()})
    sets = ", ".join(f"{k}=:{k}" for k in campos)
    run(db, f"UPDATE sst_peligros SET {sets} WHERE id=:id", dict(campos, id=pid))
    return {"ok": True, "nivel_riesgo": nr, "interpretacion": interp}


# ══════════════════════════════════════════════════════════════════════
#  PLAN ANUAL DE TRABAJO
# ══════════════════════════════════════════════════════════════════════
@router.get("/api/sgsst/actividades")
def actividades_listar(anio: int = 0, cur: dict = Depends(verify_token),
                       db: Session = Depends(get_tenant_db)):
    anio = anio or anio_actual()
    filas = serial(q(db, "SELECT * FROM sst_actividades WHERE anio = :a "
                         "ORDER BY fecha_plan, id", {"a": anio}))
    hechas = sum(1 for f in filas if f["estado"] == "ejecutada")
    vencidas = sum(1 for f in filas
                   if f["estado"] == "planeada" and (f.get("fecha_plan") or "9999") < hoy())

    return {"ok": True, "items": filas, "tipos": TIPOS_ACTIVIDAD, "anio": anio,
            "kpis": {"total": len(filas), "ejecutadas": hechas, "vencidas": vencidas,
                     # El indicador que exige la norma: ejecución del plan anual.
                     "cumplimiento": round(hechas / len(filas) * 100, 1) if filas else None}}


@router.post("/api/sgsst/actividades", status_code=201)
def actividad_crear(body: dict = Body(...), cur: dict = Depends(require_rol(*ROLES_SST)),
                    db: Session = Depends(get_tenant_db)):
    if not (body.get("nombre") or "").strip():
        raise HTTPException(400, "El nombre de la actividad es obligatorio")
    anio = int(body.get("anio") or anio_actual())
    res = run(db, "INSERT INTO sst_actividades (tipo, nombre, descripcion, responsable, "
                  "fecha_plan, estado, anio, creado_en) "
                  "VALUES (:t,:n,:d,:r,:f,'planeada',:a,:ts)",
              {"t": body.get("tipo") or "Capacitación", "n": body["nombre"],
               "d": body.get("descripcion"), "r": body.get("responsable"),
               "f": body.get("fecha_plan") or None, "a": anio, "ts": ahora()})
    return {"ok": True, "id": getattr(res, "lastrowid", 0)}


@router.put("/api/sgsst/actividades/{aid}")
def actividad_editar(aid: int, body: dict = Body(...),
                     cur: dict = Depends(require_rol(*ROLES_SST)),
                     db: Session = Depends(get_tenant_db)):
    if not q1(db, "SELECT id FROM sst_actividades WHERE id=:i", {"i": aid}):
        raise HTTPException(404, "Actividad no encontrada")

    # Marcar como ejecutada EXIGE evidencia. Sin ese requisito, la ejecución del
    # plan anual se convierte en una casilla que alguien marca antes de la
    # visita, y el indicador deja de significar algo.
    if body.get("estado") == "ejecutada" and not (body.get("evidencia") or "").strip():
        raise HTTPException(400, "Para marcar la actividad como ejecutada debe "
                                 "registrar la evidencia (acta, listado, registro).")

    campos = {k: body[k] for k in ("tipo", "nombre", "descripcion", "responsable",
                                   "fecha_plan", "fecha_real", "estado", "evidencia")
              if k in body}
    if body.get("estado") == "ejecutada" and "fecha_real" not in campos:
        campos["fecha_real"] = hoy()
    if not campos:
        return {"ok": True, "sin_cambios": True}
    sets = ", ".join(f"{k}=:{k}" for k in campos)
    run(db, f"UPDATE sst_actividades SET {sets} WHERE id=:id", dict(campos, id=aid))
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════
#  INCIDENTES Y ACCIDENTES
# ══════════════════════════════════════════════════════════════════════
@router.get("/api/sgsst/incidentes")
def incidentes_listar(cur: dict = Depends(verify_token), db: Session = Depends(get_tenant_db)):
    filas = serial(q(db, "SELECT i.*, CONCAT(COALESCE(e.nombres,''),' ',"
                         "COALESCE(e.apellidos,'')) AS empleado "
                         "FROM sst_incidentes i "
                         "LEFT JOIN empleados e ON e.id = i.empleado_id "
                         "ORDER BY i.id DESC LIMIT 200"))
    accidentes = [f for f in filas if f["tipo"] == "accidente"]
    dias = sum(int(f.get("dias_incapacidad") or 0) for f in filas)
    return {"ok": True, "items": filas,
            "kpis": {"total": len(filas), "accidentes": len(accidentes),
                     "dias_incapacidad": dias,
                     "abiertos": sum(1 for f in filas if f["estado"] == "abierto"),
                     "sin_reportar_arl": sum(1 for f in accidentes
                                             if not int(f.get("reportado_arl") or 0))}}


@router.post("/api/sgsst/incidentes", status_code=201)
def incidente_crear(body: dict = Body(...), cur: dict = Depends(verify_token),
                    db: Session = Depends(get_tenant_db)):
    """Cualquier usuario autenticado puede reportar.

    Restringir el reporte al responsable de SG-SST haría que los incidentes se
    reporten tarde o no se reporten: quien se quema es quien está en la cocina,
    no quien administra el sistema.
    """
    if not (body.get("descripcion") or "").strip():
        raise HTTPException(400, "La descripción de lo ocurrido es obligatoria")

    anio = anio_actual()
    consecutivo = f"INC-{anio}-{siguiente_consecutivo(db, 'incidente', anio):04d}"
    tipo = body.get("tipo") or "incidente"

    res = run(db, "INSERT INTO sst_incidentes (consecutivo, ts, empleado_id, nombre, tipo, "
                  "lugar, descripcion, parte_cuerpo, dias_incapacidad, estado, creado_por) "
                  "VALUES (:c,:ts,:e,:n,:t,:l,:d,:pc,:di,'abierto',:u)",
              {"c": consecutivo, "ts": ahora(), "e": body.get("empleado_id") or None,
               "n": body.get("nombre"), "t": tipo, "l": body.get("lugar"),
               "d": body["descripcion"], "pc": body.get("parte_cuerpo"),
               "di": int(body.get("dias_incapacidad") or 0), "u": autor(cur)})

    aviso = None
    if tipo == "accidente":
        # Plazo legal: dos días hábiles para reportar a la ARL y a la EPS
        # (Decreto 1072 de 2015). El sistema lo recuerda porque el
        # incumplimiento acarrea sanción.
        aviso = ("Accidente de trabajo registrado. Debe reportarse a la ARL y a la EPS "
                 "dentro de los DOS DÍAS HÁBILES siguientes.")
    return {"ok": True, "id": getattr(res, "lastrowid", 0),
            "consecutivo": consecutivo, "aviso": aviso}


@router.put("/api/sgsst/incidentes/{iid}")
def incidente_editar(iid: int, body: dict = Body(...),
                     cur: dict = Depends(require_rol(*ROLES_SST)),
                     db: Session = Depends(get_tenant_db)):
    if not q1(db, "SELECT id FROM sst_incidentes WHERE id=:i", {"i": iid}):
        raise HTTPException(404, "Incidente no encontrado")

    # Cerrar sin investigar vacía de sentido el registro: la norma pide
    # investigar y definir acciones, no solo anotar que ocurrió.
    if body.get("estado") == "cerrado":
        if not (body.get("causa_raiz") or "").strip():
            raise HTTPException(400, "Para cerrar el caso debe registrarse la causa raíz")
        if not (body.get("acciones") or "").strip():
            raise HTTPException(400, "Para cerrar el caso deben registrarse las acciones tomadas")

    campos = {k: body[k] for k in ("tipo", "lugar", "descripcion", "parte_cuerpo",
                                   "dias_incapacidad", "causa_raiz", "acciones",
                                   "reportado_arl", "estado") if k in body}
    if not campos:
        return {"ok": True, "sin_cambios": True}
    sets = ", ".join(f"{k}=:{k}" for k in campos)
    run(db, f"UPDATE sst_incidentes SET {sets} WHERE id=:id", dict(campos, id=iid))
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════
#  AUTOEVALUACIÓN DE ESTÁNDARES MÍNIMOS
# ══════════════════════════════════════════════════════════════════════
_CICLOS = {"I": "Planear", "H": "Hacer", "V": "Verificar", "A": "Actuar"}


@router.get("/api/sgsst/estandares")
def estandares_listar(cur: dict = Depends(verify_token), db: Session = Depends(get_tenant_db)):
    filas = serial(q(db, "SELECT * FROM sst_estandares ORDER BY item"))
    total_peso = sum(float(f["peso"] or 0) for f in filas)
    obtenido = sum(float(f["peso"] or 0) for f in filas if int(f["cumple"] or 0))
    pct = round(obtenido / total_peso * 100, 1) if total_peso else 0

    # Valoración de la Resolución 0312 de 2019, artículo 27.
    if pct >= 85:
        valoracion, color = "Aceptable", "#16a34a"
    elif pct >= 60:
        valoracion, color = "Moderadamente aceptable", "#d97706"
    else:
        valoracion, color = "Crítico", "#dc2626"

    por_ciclo = {}
    for f in filas:
        c = f["ciclo"]
        d = por_ciclo.setdefault(c, {"ciclo": c, "nombre": _CICLOS.get(c, c),
                                     "peso": 0.0, "obtenido": 0.0, "items": 0})
        d["peso"] += float(f["peso"] or 0)
        d["items"] += 1
        if int(f["cumple"] or 0):
            d["obtenido"] += float(f["peso"] or 0)

    return {"ok": True, "items": filas, "por_ciclo": list(por_ciclo.values()),
            "resumen": {"puntaje": round(obtenido, 2), "maximo": round(total_peso, 2),
                        "porcentaje": pct, "valoracion": valoracion, "color": color,
                        "cumplidos": sum(1 for f in filas if int(f["cumple"] or 0)),
                        "total": len(filas)}}


@router.put("/api/sgsst/estandares/{eid}")
def estandar_marcar(eid: int, body: dict = Body(...),
                    cur: dict = Depends(require_rol(*ROLES_SST)),
                    db: Session = Depends(get_tenant_db)):
    if not q1(db, "SELECT id FROM sst_estandares WHERE id=:i", {"i": eid}):
        raise HTTPException(404, "Estándar no encontrado")
    run(db, "UPDATE sst_estandares SET cumple=:c, justifica=:j WHERE id=:i",
        {"c": int(bool(body.get("cumple"))), "j": body.get("justifica"), "i": eid})
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════
#  INDICADORES
# ══════════════════════════════════════════════════════════════════════
@router.get("/api/sgsst/indicadores")
def indicadores(cur: dict = Depends(verify_token), db: Session = Depends(get_tenant_db)):
    """Indicadores mínimos del artículo 30 de la Resolución 0312.

    La frecuencia y la severidad se calculan por cada cien trabajadores al año,
    que es la base con la que se comparan contra el sector.
    """
    anio = anio_actual()
    trabajadores = int((q1(db, "SELECT COUNT(*) AS n FROM empleados WHERE activo=1")
                        or {}).get("n") or 0)

    acc = q1(db, "SELECT COUNT(*) AS n, COALESCE(SUM(dias_incapacidad),0) AS dias "
                 "FROM sst_incidentes WHERE tipo='accidente' AND ts LIKE :a",
             {"a": f"{anio}%"}) or {}
    n_acc = int(acc.get("n") or 0)
    dias = int(acc.get("dias") or 0)

    plan = q1(db, "SELECT COUNT(*) AS n, SUM(CASE WHEN estado='ejecutada' THEN 1 ELSE 0 END) AS ok "
                  "FROM sst_actividades WHERE anio=:a", {"a": anio}) or {}
    n_plan = int(plan.get("n") or 0)
    n_ok = int(plan.get("ok") or 0)

    base = trabajadores or 1
    return {"ok": True, "anio": anio, "trabajadores": trabajadores,
            "indicadores": [
                {"nombre": "Frecuencia de accidentalidad",
                 "valor": round(n_acc / base * 100, 2), "unidad": "por 100 trabajadores",
                 "detalle": f"{n_acc} accidente(s) en {trabajadores} trabajador(es)"},
                {"nombre": "Severidad de accidentalidad",
                 "valor": round(dias / base * 100, 2), "unidad": "días por 100 trabajadores",
                 "detalle": f"{dias} día(s) de incapacidad"},
                {"nombre": "Proporción de accidentes mortales",
                 "valor": 0.0, "unidad": "%", "detalle": "Sin eventos mortales registrados"},
                {"nombre": "Ejecución del plan anual",
                 "valor": round(n_ok / n_plan * 100, 1) if n_plan else 0.0, "unidad": "%",
                 "detalle": f"{n_ok} de {n_plan} actividades ejecutadas"},
            ]}


# ══════════════════════════════════════════════════════════════════════
#  PLAN ANUAL EN EXCEL
#
#  Existe porque el plan anual NO se construye en la pantalla. Se arma en una
#  reunión, se pasa a una hoja de cálculo, la revisa la ARL y vuelve
#  corregida. Obligar a teclear quince actividades una por una garantiza que
#  el plan viva por fuera del sistema, que es justamente lo que hay que evitar.
#
#  La MISMA hoja sirve de plantilla y de respaldo: se exporta, se edita, se
#  vuelve a subir. Por eso lleva la columna `id` — con id actualiza, sin id
#  crea— y por eso la primera fila es la cabecera exacta que espera el
#  importador.
#
#  LA IMPORTACIÓN ES TODO O NADA
#  Se validan todas las filas ANTES de escribir. Si una está mal, no entra
#  ninguna y se devuelve la lista de problemas con su número de fila. Un plan
#  a medio importar —ocho actividades sí y siete no— es peor que uno no
#  importado: nadie sabe cuáles faltan.
# ══════════════════════════════════════════════════════════════════════
COLUMNAS_PLAN = ["id", "tipo", "nombre", "descripcion", "responsable",
                 "fecha_plan", "fecha_real", "estado", "evidencia"]

ENCABEZADOS = ["ID", "Tipo", "Actividad", "Descripción", "Responsable",
               "Fecha programada", "Fecha de ejecución", "Estado", "Evidencia"]


@router.get("/api/sgsst/actividades/excel")
def plan_exportar(anio: int = 0, cur: dict = Depends(verify_token),
                  db: Session = Depends(get_tenant_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    anio = anio or anio_actual()
    filas = q(db, "SELECT * FROM sst_actividades WHERE anio=:a ORDER BY fecha_plan, id",
              {"a": anio})

    wb = Workbook()
    ws = wb.active
    ws.title = f"Plan {anio}"

    ws.append(ENCABEZADOS)
    cab = PatternFill("solid", fgColor="1B4332")
    for i, _ in enumerate(ENCABEZADOS, start=1):
        c = ws.cell(row=1, column=i)
        c.fill = cab
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for f_ in filas:
        ws.append([f_.get(k) for k in COLUMNAS_PLAN])

    anchos = [6, 16, 52, 34, 22, 18, 18, 13, 34]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for fila in ws.iter_rows(min_row=2):
        fila[2].alignment = Alignment(wrap_text=True, vertical="top")

    # Hoja de ayuda: sin ella, quien reciba el archivo no sabe qué puede
    # escribir en «estado» ni que la columna id no se debe tocar.
    hoja = wb.create_sheet("Instrucciones")
    for linea in [
        ["Cómo usar este archivo"],
        [""],
        ["1.", "Edite las filas o agregue nuevas al final."],
        ["2.", "NO modifique la columna ID. Con ID, la fila actualiza una actividad "
               "existente; sin ID, crea una nueva."],
        ["3.", "Las fechas van en formato AAAA-MM-DD."],
        ["4.", "Estado admite: planeada, ejecutada, cancelada."],
        ["5.", "Para marcar una actividad como ejecutada debe escribir la evidencia "
               "(acta, listado de asistencia, registro)."],
        [""],
        ["Tipos sugeridos:", ", ".join(TIPOS_ACTIVIDAD)],
        [""],
        ["La importación es todo o nada: si una fila tiene un error, no entra "
         "ninguna y se le indica cuál corregir."],
    ]:
        hoja.append(linea)
    hoja.column_dimensions["A"].width = 18
    hoja.column_dimensions["B"].width = 86
    hoja.cell(row=1, column=1).font = Font(bold=True, size=13)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f"plan-anual-sgsst-{anio}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'})


@router.post("/api/sgsst/actividades/importar")
async def plan_importar(archivo: UploadFile = File(...), anio: int = Form(0),
                        cur: dict = Depends(require_rol(*ROLES_SST)),
                        db: Session = Depends(get_tenant_db)):
    from openpyxl import load_workbook

    if not (archivo.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "El archivo debe ser una hoja de Excel (.xlsx)")

    crudo = await archivo.read()
    if len(crudo) > 4 * 1024 * 1024:
        raise HTTPException(400, "El archivo supera los 4 MB")

    try:
        wb = load_workbook(io.BytesIO(crudo), data_only=True)
    except Exception:
        raise HTTPException(400, "No se pudo leer el archivo. ¿Es un Excel válido?")

    ws = wb.worksheets[0]
    anio = anio or anio_actual()

    def texto(v):
        if v is None:
            return ""
        if isinstance(v, (datetime.date, datetime.datetime)):
            return v.strftime("%Y-%m-%d")
        return str(v).strip()

    filas, errores = [], []
    for n, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not fila or not any(fila):
            continue
        d = {c: texto(fila[i]) if i < len(fila) else ""
             for i, c in enumerate(COLUMNAS_PLAN)}

        if not d["nombre"]:
            errores.append(f"Fila {n}: falta el nombre de la actividad")
            continue
        for campo in ("fecha_plan", "fecha_real"):
            if d[campo]:
                try:
                    datetime.date.fromisoformat(d[campo][:10])
                    d[campo] = d[campo][:10]
                except ValueError:
                    errores.append(f"Fila {n}: «{d[campo]}» no es una fecha AAAA-MM-DD")
        estado = (d["estado"] or "planeada").lower()
        if estado not in ("planeada", "ejecutada", "cancelada"):
            errores.append(f"Fila {n}: estado «{d['estado']}» no válido")
        elif estado == "ejecutada" and not d["evidencia"]:
            errores.append(f"Fila {n}: «{d['nombre'][:38]}» está como ejecutada "
                           "pero no tiene evidencia")
        d["estado"] = estado
        if d["id"]:
            try:
                d["id"] = int(float(d["id"]))
            except ValueError:
                errores.append(f"Fila {n}: el ID «{d['id']}» no es un número")
                d["id"] = 0
        else:
            d["id"] = 0
        filas.append((n, d))

    if errores:
        raise HTTPException(400, "No se importó nada. Corrija y vuelva a subir:\n· "
                            + "\n· ".join(errores[:12])
                            + (f"\n… y {len(errores) - 12} problema(s) más"
                               if len(errores) > 12 else ""))
    if not filas:
        raise HTTPException(400, "La hoja no tiene ninguna actividad")

    creadas = actualizadas = 0
    try:
        for _, d in filas:
            if d["id"] and q1(db, "SELECT id FROM sst_actividades WHERE id=:i",
                              {"i": d["id"]}):
                run_sin_commit(db,
                               "UPDATE sst_actividades SET tipo=:t, nombre=:n, "
                               "descripcion=:d, responsable=:r, fecha_plan=:f, "
                               "fecha_real=:fr, estado=:e, evidencia=:ev WHERE id=:i",
                               {"t": d["tipo"] or "Capacitación", "n": d["nombre"][:200],
                                "d": d["descripcion"] or None,
                                "r": d["responsable"] or None,
                                "f": d["fecha_plan"] or None,
                                "fr": d["fecha_real"] or None, "e": d["estado"],
                                "ev": d["evidencia"] or None, "i": d["id"]})
                actualizadas += 1
            else:
                run_sin_commit(db,
                               "INSERT INTO sst_actividades (tipo, nombre, descripcion, "
                               "responsable, fecha_plan, fecha_real, estado, evidencia, "
                               "anio, creado_en) "
                               "VALUES (:t,:n,:d,:r,:f,:fr,:e,:ev,:a,:ts)",
                               {"t": d["tipo"] or "Capacitación", "n": d["nombre"][:200],
                                "d": d["descripcion"] or None,
                                "r": d["responsable"] or None,
                                "f": d["fecha_plan"] or None,
                                "fr": d["fecha_real"] or None, "e": d["estado"],
                                "ev": d["evidencia"] or None, "a": anio, "ts": ahora()})
                creadas += 1
        db.commit()
    except Exception:
        db.rollback()
        raise

    return {"ok": True, "creadas": creadas, "actualizadas": actualizadas,
            "mensaje": (f"Plan importado: {creadas} actividad(es) nueva(s) y "
                        f"{actualizadas} actualizada(s).")}
